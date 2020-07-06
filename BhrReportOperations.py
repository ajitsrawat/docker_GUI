import openpyxl
import os
import datetime
#import pythoncom
#import win32com.client as win32
#from openpyxl import workbook
from openpyxl.styles import Font, Fill, Border, Side, NamedStyle, PatternFill, Alignment
from openpyxl.styles.colors import YELLOW, RED, GREEN, BLUE
from openpyxl.chart import LineChart, BarChart, BarChart3D, Series, Reference, AreaChart3D
from jira import JIRA

blockedAreaCountR20_2=0


def deriveCellIndex(queryIndex):
    if(queryIndex == 0):
        return(3,2)
    
    if(queryIndex == 1):
        return(4,2)

    if(queryIndex == 2):
        return(4,7)
    
    if(queryIndex == 3):
        return(3,7)

    if(queryIndex == 4):
        return(4,8)

    if(queryIndex == 5):
        return(3,8)

    if(queryIndex == 6):
        return(5,3)

    if(queryIndex == 7):
        return(6,3)

    if(queryIndex == 8):
        return(7,3)

    if(queryIndex == 9):
        return(8,8)

    if(queryIndex == 10):
        return(3,9)

    if(queryIndex == 11):
        return(8,9)

    if(queryIndex == 12):
        return(4,9)

    if(queryIndex == 13):
        return(5,4)

    if(queryIndex == 14):
        return(3,12)

    if(queryIndex == 15):
        return(4,12)

    if(queryIndex == 16):
        return(3,10)

    if(queryIndex == 17):
        return(4,10)

    if(queryIndex == 18):
        return(3,6)

    if(queryIndex == 19):
        return(4,6)

    if(queryIndex == 20):
        return(3,11)

    if(queryIndex == 21):
        return(4,11)

    if(queryIndex == 22):
        return(8,11)

    if(queryIndex == 23):
        return(10,2)

    if(queryIndex == 24):
        return(10,3)

    if(queryIndex == 25):
        return(10,6)

    if(queryIndex == 26):
        return(11,2)

    if(queryIndex == 27):
        return(12,2)
    
    if(queryIndex == 28):
        return(13,2)

    if(queryIndex == 29):
        return(10,4)

    if(queryIndex == 30):
        return(10,5)
    
#Populate Query List from Excel File
def readQueries(querySheet):
    queryList =[]
    for i in range(2, querySheet.max_row+1):
        if(querySheet.cell(i, 1).value != None):
            queryList.append(querySheet.cell(i, 1).value)
    listSize = len(queryList)
    for i in range(0, listSize):
        print(queryList[i])

    return queryList  


### Find if bug is already added to list of issues
def isUniqueIssue(allBugList, key):
    print(allBugList)
    for bugId in allBugList:
        print("bug Id =", bugId)
        if(str(bugId) == str(key)):##issue already in list
            print("Bug already in list hence return")
            return 0

    print("this is unique added to bugList")
    allBugList.append(key)
    return 1 

#Find the Bug Id along with it's criticality. 
def findAndFillFaultId(release, jira, issue, detailsSheet, queryIndex, allBugList):
    bugStr =''
    for issueLinked in issue.fields.issuelinks:
        print("Going into Loop for", issue.key)
        issue1 = jira.issue_link(issueLinked)
        issue2 = jira.issue(issue1.inwardIssue.key)
        issue3 = jira.issue(issue1.outwardIssue.key)
        #print(issue.fields.affectedVersion)


        if(str(issue2.fields.issuetype) == 'Bug'):
            if( (isUniqueIssue(allBugList, issue2.key) ==1) and (str(issue2.fields.status) != 'Closed')):
                print("Adding id", issue2.key)
                bugStr = bugStr + str(issue2.key)+' '+ str(issue2.fields.priority)+ ' '+ str(issue2.fields.status)+' '+ str(issue2.fields.created.rsplit('T')[0])+' ' + str(issue2.fields.summary)+ '\n'
                print(bugStr)

        if(str(issue3.fields.issuetype) =='Bug'):
            if( (isUniqueIssue(allBugList, issue3.key) ==1) and (str(issue3.fields.status) != 'Closed')):
                print("Adding id", issue3.key)
                bugStr = bugStr + str(issue3.key)+' '+ str(issue3.fields.priority)+' '+ str(issue3.fields.status)+ ' '+ str(issue3.fields.created.rsplit('T')[0])+' ' + str(issue3.fields.summary)+ '\n'
                print(bugStr)

    detailsSheet.cell(queryIndex+2,5).value = str(detailsSheet.cell(queryIndex+2,5).value) + str(bugStr)



#Update Health Report for all blocked issues. 
def updateHealthReport(release, jira, issueList, healthReportSheet,detailsSheet, queryIndex, issues_Planned,issues_Done):
    global blockedAreaCountR20_2
    ustIdList = ' '
    allBugList = []

    ### Fill UST count in Health Report Mark it Red
    cellIndex = deriveCellIndex(queryIndex)
    #print(cellIndex[0]," ", cellIndex[1])
    blockedIssueCount = len(issueList)
    healthReportSheet.cell(cellIndex[0], cellIndex[1]).value = "Planned USTs = " + str(issues_Planned) +'\n Blocked UST = ' + str(blockedIssueCount) +  '\n Done UST = ' + str(issues_Done)
    if(blockedIssueCount> 0):# Fill Red Color for Blocked Area
        healthReportSheet.cell(cellIndex[0], cellIndex[1]).fill = PatternFill(fgColor='00FF0000', fill_type = 'solid')
    elif(issues_Done/issues_Planned < 0.3): # Fill Yellow to show low Count of Done
        healthReportSheet.cell(cellIndex[0], cellIndex[1]).fill = PatternFill(fgColor='00FFFFFF', fill_type = 'solid')
    elif(issues_Done/issues_Planned > 0.7): # Fill Green to show progress on Done
        healthReportSheet.cell(cellIndex[0], cellIndex[1]).fill = PatternFill(fgColor='0000FF00', fill_type = 'solid')
    elif(issues_Done/issues_Planned <=0.7): # Fill Dark Green for high number of Done
        healthReportSheet.cell(cellIndex[0], cellIndex[1]).fill = PatternFill(fgColor='00FFFF00', fill_type = 'solid')



    print("Blocked Areas wise UST count", blockedIssueCount)


    #Fill Details View containing Ids of all USTs
    for i in range(0, blockedIssueCount):
        ustIdList = str(issueList[i].key) + ", "+ ustIdList
    print(ustIdList)
    if(detailsSheet.cell(queryIndex+2,4).value == None):
        detailsSheet.cell(queryIndex+2,4).value = ustIdList
    else:
        detailsSheet.cell(queryIndex+2,4).value= detailsSheet.cell(queryIndex+2,4).value + ustIdList

    ## Fill Jira details (related to Blocked USTs)
    for issue in issueList:
        findAndFillFaultId(release, jira, issue, detailsSheet, queryIndex, allBugList)
    if(release == "R20.8"):
        blockedAreaCountR20_2 = blockedAreaCountR20_2 +1




# def sendHealthReportToTeam():
#     pythoncom.CoInitialize()
#     outlook = win32.Dispatch('outlook.application')
#     mail = outlook.CreateItem(0)
#     # mail.To = 'ajit.rawat@nokia.com'
#     mail.To = 'ajit.rawat@nokia.com; daloka.reddy@nokia.com; shyam.c@nokia.com; deepak.1.toshniwal@nokia.com; vijayakumar.vempadapu@nokia.com; balaji.venkataraman@nokia.com; mohanraj.s@nokia.com; kalyan.pingali@nokia.com; karthikeyan.1.krishnan@nokia.com; senthil.1.sundaram@nokia.com; krishna.pr@nokia.com; pradeep.c_p@nokia.com; sakti.sahoo@nokia.com;deva.baswa@nokia.com; chaitanya.yemineni@nokia.com; rakesh.kumar_t@nokia.com; prabhakaran.t@nokia.com'
#     mail.Subject = "Basic Health Report"
#     mail.Attachments.Add("C:/ASRAWAT/test/BasicHealthReport/BasicHealthForDay2DayWork.xlsx")
#     ##
#     ##    #mail.HTMLBody = "<html><body>Overall System Health <img src=""cid:MyId1""></body></html>"
#     ##    mail.HTMLBody = "<html><body><h2>Overall System Health <br><img src=""cid:MyId1"" width = ""500"" height= ""400""></h2><h2><br>Overall System Trend<br> <img src=""cid:MyId2"" width = ""500"" height= ""400""></h2><h2><br>HLR Trend<br> <img src=""cid:MyId3"" width = ""500"" height= ""400""></h2><h2><br>HSS Trend<br> <img src=""cid:MyId4"" width = ""500"" height= ""400""></h2><h2><br>UDM Trend<br> <img src=""cid:MyId5"" width = ""500"" height= ""400""></h2></body></html>"
#     print("Sending Mail")
#     mail.Send()


# def sendHealthReportToMeOnly():
#     pythoncom.CoInitialize()
#     outlook = win32.Dispatch('outlook.application')
#     mail = outlook.CreateItem(0)
#     mail.To = 'ajit.rawat@nokia.com'
#     mail.Subject = "Basic Health Report"
#     # mail.To = 'ajit.rawat@nokia.com; daloka.reddy@nokia.com; shyam.c@nokia.com; deepak.1.toshniwal@nokia.com; vijayakumar.vempadapu@nokia.com; balaji.venkataraman@nokia.com; mohanraj.s@nokia.com; kalyan.pingali@nokia.com; karthikeyan.1.krishnan@nokia.com; senthil.1.sundaram@nokia.com; krishna.pr@nokia.com; pradeep.c_p@nokia.com; sakti.sahoo@nokia.com;deva.baswa@nokia.com; chaitanya.yemineni@nokia.com; rakesh.kumar_t@nokia.com; prabhakaran.t@nokia.com'
#     currentWk = datetime.date.today().isocalendar()[1]
#     mail.Attachments.Add("C:/ASRAWAT/test/BasicHealthReport/BasicHealthForDay2DayWork_wk'+str(currentWk)+'.xlsx")
#     ##
#     ##    #mail.HTMLBody = "<html><body>Overall System Health <img src=""cid:MyId1""></body></html>"
#     ##    mail.HTMLBody = "<html><body><h2>Overall System Health <br><img src=""cid:MyId1"" width = ""500"" height= ""400""></h2><h2><br>Overall System Trend<br> <img src=""cid:MyId2"" width = ""500"" height= ""400""></h2><h2><br>HLR Trend<br> <img src=""cid:MyId3"" width = ""500"" height= ""400""></h2><h2><br>HSS Trend<br> <img src=""cid:MyId4"" width = ""500"" height= ""400""></h2><h2><br>UDM Trend<br> <img src=""cid:MyId5"" width = ""500"" height= ""400""></h2></body></html>"
#     print("Sending Mail")
#     mail.Send()


def generateWeeklyTrend(jira, weeklyTrendSheet,graphSheet):
    print("Trend")
    global blockedAreaCountR18_5SP6
    global blockedAreaCountR20_2
    blokcedUSTQueryR20 = 'project in (RGSOL) AND issuetype = "User Story" AND fixVersion = 2019PI4  AND affectedVersion in ("Nokia Registers 20") AND status = Blocked'
    blokcedUSTQueryR185MP6 = 'project in (RGSOL) AND issuetype = "User Story" AND fixVersion = 2019PI4 AND affectedVersion in ("Nokia Registers 18.5MP6") AND status = Blocked'
    issueListR20 = jira.search_issues(blokcedUSTQueryR20)
    print(len(issueListR20))
    issueListR185MP6 = jira.search_issues(blokcedUSTQueryR185MP6)

    currentWk = datetime.date.today().isocalendar()[1]
    for i in range(3, weeklyTrendSheet.max_row):
        if(weeklyTrendSheet.cell(i, 1).value is None):
            weeklyTrendSheet.cell(i, 1).value = currentWk
            weeklyTrendSheet.cell(i, 3).value = blockedAreaCountR20_2
            weeklyTrendSheet.cell(i, 2).value = len(issueListR20)
            # Copy system area values from copied cells and then delete those cells.
            weeklyTrendSheet.cell(i, 4).value = weeklyTrendSheet.cell(20, 4).value
            weeklyTrendSheet.cell(i, 5).value = weeklyTrendSheet.cell(20, 5).value
            weeklyTrendSheet.cell(i, 6).value =weeklyTrendSheet.cell(20, 6).value
            weeklyTrendSheet.cell(i, 7).value =weeklyTrendSheet.cell(20, 7).value
            weeklyTrendSheet.cell(i, 8).value = weeklyTrendSheet.cell(20, 8).value
            weeklyTrendSheet.cell(i, 9).value = weeklyTrendSheet.cell(20, 9).value
            ### For 18.5 MP6
            weeklyTrendSheet.cell(i, 12).value = currentWk
            weeklyTrendSheet.cell(i, 14).value = blockedAreaCountR18_5SP6
            weeklyTrendSheet.cell(i, 13).value = len(issueListR185MP6)

            break
        if(int(weeklyTrendSheet.cell(i, 1).value)  < currentWk):
            continue
        if (int(weeklyTrendSheet.cell(i, 1).value) == currentWk):# Entry already present
            break
    generateHealthTrendGraph(weeklyTrendSheet, graphSheet)


def generateHealthTrendGraph(weeklyTrendSheet, graphSheet):
    print("Drawing Overall graph")
    chart_health_overall = BarChart()
    chart_health_overall.type = "col"
    chart_health_overall.name = " Overall Health Trend"
    #chart_health_overall.style = 13
    chart_health_overall.title = "Overall Health Trend"
    chart_health_overall.y_axis.title = 'System Area'
    chart_health_overall.x_axis.title = 'WeeklyProgress'
    data = Reference(weeklyTrendSheet, min_col=2, min_row=2, max_row=15, max_col=3)
    cats = Reference(weeklyTrendSheet, min_col=1, min_row=2, max_row=15)
    chart_health_overall.add_data(data, titles_from_data=True)
    chart_health_overall.set_categories(cats)
    #chart_health_overall.shape = 30
    graphSheet.add_chart(chart_health_overall,"B2")

# For Security
    print("Drawing Security graph")
    chart_health_Security = BarChart()
    chart_health_Security.type = "col"
    chart_health_Security.name = " Security Area Trends"
    #chart_health_Security.style = 13
    chart_health_Security.title = "Security Area Trend"
    chart_health_Security.y_axis.title = 'Security Area'
    chart_health_Security.x_axis.title = 'WeeklyProgress'
    data = Reference(weeklyTrendSheet, min_col=6, min_row=2, max_row=15, max_col=6)
    cats = Reference(weeklyTrendSheet, min_col=1, min_row=2, max_row=15)
    chart_health_Security.add_data(data, titles_from_data=True)
    chart_health_Security.set_categories(cats)
    #chart_health_Security.shape = 30
    graphSheet.add_chart(chart_health_Security,"M2")

    # For Serviceability
    print("Drawing Serviceability graph")
    chart_health_Serviceability = BarChart()
    chart_health_Serviceability.type = "col"
    chart_health_Serviceability.name = " Serviceability Area Trends"
    # chart_health_Serviceability.style = 13
    chart_health_Serviceability.title = "Serviceability Area Trend"
    chart_health_Serviceability.y_axis.title = 'Serviceability Area'
    chart_health_Serviceability.x_axis.title = 'WeeklyProgress'
    data = Reference(weeklyTrendSheet, min_col=5, min_row=2, max_row=15, max_col=5)
    cats = Reference(weeklyTrendSheet, min_col=1, min_row=2, max_row=15)
    chart_health_Serviceability.add_data(data, titles_from_data=True)
    chart_health_Serviceability.set_categories(cats)
    # chart_health_Serviceability.shape = 30
    graphSheet.add_chart(chart_health_Serviceability, "B20")

    # For 5G
    print("Drawing 5G graph")
    chart_health_5G = BarChart()
    chart_health_5G.type = "col"
    chart_health_5G.name = " 5G Area Trends"
    # chart_health_5G.style = 13
    chart_health_5G.title = "5G Area Trend"
    chart_health_5G.y_axis.title = '5G Area'
    chart_health_5G.x_axis.title = 'WeeklyProgress'
    data = Reference(weeklyTrendSheet, min_col=9, min_row=2, max_row=15, max_col=9)
    cats = Reference(weeklyTrendSheet, min_col=1, min_row=2, max_row=15)
    chart_health_5G.add_data(data, titles_from_data=True)
    chart_health_5G.set_categories(cats)
    # chart_health_5G.shape = 30
    graphSheet.add_chart(chart_health_5G, "B60")

    # For RET&MAV
    print("Drawing RET&MAV graph")
    chart_health_RETMAV = BarChart()
    chart_health_RETMAV.type = "col"
    chart_health_RETMAV.name = " RET&MAV Area Trends"
    # chart_health_RETMAV.style = 13
    chart_health_RETMAV.title = "RET&MAV Area Trend"
    chart_health_RETMAV.y_axis.title = 'RET&MAV Area'
    chart_health_RETMAV.x_axis.title = 'WeeklyProgress'
    data = Reference(weeklyTrendSheet, min_col=7, min_row=2, max_row=15, max_col=7)
    cats = Reference(weeklyTrendSheet, min_col=1, min_row=2, max_row=15)
    chart_health_RETMAV.add_data(data, titles_from_data=True)
    chart_health_RETMAV.set_categories(cats)
    # chart_health_RETMAV.shape = 30
    graphSheet.add_chart(chart_health_RETMAV, "B40")

    # For NetAct
    print("Drawing NetAct graph")
    chart_health_NETACT = BarChart()
    chart_health_NETACT.type = "col"
    chart_health_NETACT.name = " NetAct Area Trends"
    # chart_health_NETACT.style = 13
    chart_health_NETACT.title = "NetAct Area Trend"
    chart_health_NETACT.y_axis.title = 'NetAct Area'
    chart_health_NETACT.x_axis.title = 'WeeklyProgress'
    data = Reference(weeklyTrendSheet, min_col=8, min_row=2, max_row=15, max_col=8)
    cats = Reference(weeklyTrendSheet, min_col=1, min_row=2, max_row=15)
    chart_health_NETACT.add_data(data, titles_from_data=True)
    chart_health_NETACT.set_categories(cats)
    # chart_health_NETACT.shape = 30
    graphSheet.add_chart(chart_health_NETACT, "M40")

    # For Performance
    print("Drawing Performance graph")
    chart_health_Perf = BarChart()
    chart_health_Perf.type = "col"
    chart_health_Perf.name = " Performance Area Trends"
    # chart_health_Perf.style = 13
    chart_health_Perf.title = "Performance Area Trend"
    chart_health_Perf.y_axis.title = 'Performance Area'
    chart_health_Perf.x_axis.title = 'WeeklyProgress'
    data = Reference(weeklyTrendSheet, min_col=4, min_row=2, max_row=15, max_col=4)
    cats = Reference(weeklyTrendSheet, min_col=1, min_row=2, max_row=15)
    chart_health_Perf.add_data(data, titles_from_data=True)
    chart_health_Perf.set_categories(cats)
    # chart_health_Perf.shape = 30
    graphSheet.add_chart(chart_health_Perf, "M20")
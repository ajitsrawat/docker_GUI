from flask import Flask, render_template
import openpyxl
import os.path, time

from flask import send_file
from flask import Response
import sys
from datetime import datetime
from time import sleep

app = Flask(__name__, template_folder= 'C:/ASRAWAT/test/Docker/GUI/template/')
value ='I am great'
import BasicHealthReportGenerator as BHR


@app.route('/count')
def stopWatch():
    def streamer():
        while True:
            yield 'a \n'
            yield 'b \n'
            sleep(1)

    return Response(streamer())

@app.route('/HealthReport_OnlineVersion')
def onlineVersion():
    #workBook = BHR.generateBasicHealthReport(2)
    workBook = openpyxl.load_workbook("C:/ASRAWAT/test/Docker/GUI/template/BasicHealthForDay2DayWork.xlsx")
    healthReportSheet = workBook["BasicHealthReport_R20.8"]
    detailsSheet = workBook["Details_R20.8"]
    list = [[],[],[],[],[],[],[],[],[],[],[], [], [], [], [], [],[],[],[],[],[],[],[],[],[],[], [], [], [], [], [],[],[],[],[],[],[],[],[],[],[], [], [], [], [], [],[],[],[],[],[],[],[],[],[],[], [], [], [], []]
    for i in range(2, 14):
        for j in range (1, 14):
            list[i-2].append(healthReportSheet.cell(i, j).value)
            print(healthReportSheet.cell(i,j).value)

    for i in range(15, 45):
        for j in range (2, 14):
            list[i-2].append(detailsSheet.cell(i-14, j).value)
            print(detailsSheet.cell(i,j).value)
    return render_template("HealthReport_1.html", li = list)
    #return (send_file('C:/ASRAWAT/test/WebServerFlask/BasicHealthForDay.xlsx', attachment_filename='BasicHealthForDay.xlsx'))
    #return "Report Generated"

# =============================================================================
# @app.route('/downloadUpdatedCopy')
# def downloadUpdatedCopy():
#     workBook = BHR.generateBasicHealthReport(1)
#     return (send_file('C:/ASRAWAT/test/Docker/GUI/template/BasicHealthForDay2DayWork.xlsx', attachment_filename='BasicHealthForDay2DayWork.xlsx'))
#     #return "Report Generated"
# =============================================================================

@app.route('/download')
def download():

    return (send_file('C:/ASRAWAT/test/Docker/GUI/template/BasicHealthForDay2DayWork.xlsx', attachment_filename='BasicHealthForDay2DayWork.xlsx'))
    #return "Report Generated"



@app.route('/home')
@app.route('/')
def home():
    file = 'C:/ASRAWAT/test/WebServerFlask/BasicHealthForDay.xlsx'
    lastModified=str(time.ctime(os.path.getmtime(file)))
    print(lastModified)
    return render_template("home.html", time = lastModified)

if __name__ == '__main__':
    app.run(debug=False)#, host = '10.143.81.229')
import openpyxl
import os
import datetime
#from openpyxl import workbook
from openpyxl.styles import Font, Fill, Border, Side, NamedStyle, PatternFill, Alignment
from openpyxl.styles.colors import YELLOW, RED, GREEN, BLUE
import BhrReportOperations as RO
from jira import JIRA
import tkinter
from tkinter import messagebox

perfUstCount = 0
servcUstCount =0
securityUstcount = 0
RET_MAVUstCount = 0
NetActUstCount = 0
fiveGUstCount = 0


def generateBasicHealthReportReleaseWise(workBook, jira, query, queryIndex, release):
    healthReportSheet_R1 = workBook["BasicHealthReport_R20.8"]
    healthReportSheet_R2 = workBook["BasicHealthReport_R19MP6"]
    #healthReportSheet_R185mp6 = workBook["BasicHealthReport_R18.5MP6"]
    issues_Planned = len(jira.search_issues(query))
    print("Planned = ", issues_Planned)
    if(issues_Planned>0):
        issues_Done = len(jira.search_issues(query + ' and status = Done '))
        print("Done = ", issues_Done)
        issueList = jira.search_issues(query+ 'and status = Blocked ')
        if(len(issueList)>=0):
            #blockedIssueCount = blockedIssueCount +  len(issueList)
            print(query)
            #print("\n total blocked issue = ", len(issueList),'\n', "in Release", release)
            if(release == 'R20.8'):
                detailsSheet = workBook["Details_R20.8"]
                RO.updateHealthReport(release, jira, issueList, healthReportSheet_R1,detailsSheet, queryIndex, issues_Planned,issues_Done )
            elif(release == "R19MP6"):
                detailsSheet = workBook["Details_R19MP6"]
                RO.updateHealthReport(release, jira, issueList, healthReportSheet_R2, detailsSheet, queryIndex, issues_Planned,issues_Done)



def resetHealthSheets(workBook):
    healthReportSheet_R1 = workBook["BasicHealthReport_R20.8"]
    healthReportSheet_R2 = workBook["BasicHealthReport_R19MP6"]
    jiraReport = workBook["JirasRaised"]
    # Reset the HealthSheet table.
    for i in range(3, 70):
        for j in range(2, 13):
            healthReportSheet_R1.cell(i, j).value = ''
            healthReportSheet_R1.cell(i, j).fill = PatternFill(fgColor='ffffffff', fill_type='solid')
            healthReportSheet_R2.cell(i, j).value = ''
            healthReportSheet_R2.cell(i, j).fill = PatternFill(fgColor='ffffffff', fill_type='solid')
            if(j<9): # JiraRaised sheet has only 7 coloumns
                jiraReport.cell(i, j-1).value = ''
            if(j ==12 and (i>4)):
                healthReportSheet_R1.cell(i, j).value = 'Implicitly Covered'

    healthReportSheet_R1.cell(15, 1).fill = PatternFill(fgColor='00FF0000', fill_type='solid')
    healthReportSheet_R1.cell(15, 1).value = "Some USTs are Blocked"
    healthReportSheet_R1.cell(16, 1).fill = PatternFill(fgColor='00FFFFFF', fill_type='solid')
    healthReportSheet_R1.cell(16, 1).value = "Less than 30% USTs Done"
    healthReportSheet_R1.cell(17, 1).fill = PatternFill(fgColor='00FFFF00', fill_type='solid')
    healthReportSheet_R1.cell(17, 1).value = "Less than 70% USTs Done"
    healthReportSheet_R1.cell(18, 1).fill = PatternFill(fgColor='0000FF00', fill_type='solid')
    healthReportSheet_R1.cell(18, 1).value = "More than 70% USTs Done"

    healthReportSheet_R2.cell(15, 1).fill = PatternFill(fgColor='00FF0000', fill_type='solid')
    healthReportSheet_R2.cell(15, 1).value = "Some USTs are Blocked"
    healthReportSheet_R2.cell(16, 1).fill = PatternFill(fgColor='00FFFFFF', fill_type='solid')
    healthReportSheet_R2.cell(16, 1).value = "Less than 30% USTs Done"
    healthReportSheet_R2.cell(17, 1).fill = PatternFill(fgColor='00FFFF00', fill_type='solid')
    healthReportSheet_R2.cell(17, 1).value = "Less than 70% USTs Done"
    healthReportSheet_R2.cell(18, 1).fill = PatternFill(fgColor='0000FF00', fill_type='solid')
    healthReportSheet_R2.cell(18, 1).value = "More than 70% USTs Done"

    #include resetting charts as well
    chartSheet = workBook["HealthTrendCharts"]
    workBook.remove(chartSheet)
    chartSheet = workBook.create_sheet("HealthTrendCharts", 6)
    chartSheet.title = "HealthTrendCharts"

def fillJiraRaisedBySyVeinReport(workBook, jira):
    jiraSheet = workBook["JirasRaised"]
    jiraQuery = 'project in (REG, IMS, ZTS, TM_SDLST7, TM_SDLST3, TM_SDLST12, ICE) AND issuetype = Bug AND (assignee in membersOf(I_REGISTERSRD_SYVETEAM) OR reporter in membersOf(I_REGISTERSRD_SYVETEAM)) AND status != Closed AND summary !~ CLONE AND created > "2020/03/11" ORDER BY created ASC'
    issueList = jira.search_issues(jiraQuery)
    row=3
    for issue in issueList:
        jiraSheet.cell(row, 1).value = str(issue.key)
        jiraSheet.cell(row, 2).value = str(issue.fields.summary)
        jiraSheet.cell(row, 3).value = str(issue.fields.status)
        jiraSheet.cell(row, 4).value = str(issue.fields.priority)
        try:
            jiraSheet.cell(row, 5).value = str(issue.fields.fixVersions.pop().name)
        except:
            print("Error in Fix version for bug ", issue.key )
        jiraSheet.cell(row, 6).value = str(issue.fields.created.rsplit('T')[0])
        jiraSheet.cell(row, 7).value = str(issue.fields.creator)
        row = row+1


# Main function (entry point function)
def generateBasicHealthReport(guiOrWeb=0):
    # global perfUstCount
    # global servcUstCount
    # global securityUstcount
    # global RET_MAVUstCount
    # global NetActUstCount
    # global fiveGUstCount
    # global perfUstCount
    currentWk = datetime.date.today().isocalendar()[1]
    try:
        workBook = openpyxl.load_workbook('C:/ASRAWAT/test/Docker/GUI/template/BasicHealthForDay2DayWork.xlsx')
    except FileNotFoundError:
        print("MyError: New Week Started copying previous week file to new one.")
        #workBook = openpyxl.load_workbook('C:/ASRAWAT/test/BasicHealthReport/BasicHealthForDay2DayWork_wk' + str(currentWk-1) + '.xlsx')
    print(workBook.sheetnames) # all names

    if(guiOrWeb == 2 ):
        return workBook

    weeklyTrendSheet = workBook["Trends"]
    querySheet = workBook["Details_R20.8"]
    r19MP6Sheet = workBook["Details_R19MP6"]
    # Reset the QuerySheet table.
    queryList = RO.readQueries(querySheet)
    listSize = len(queryList)
    for i in range(2, listSize + 2):
        for j in range(4, 11):
            querySheet.cell(i, j).value = ''
            r19MP6Sheet.cell(i, j).value = ''

    resetHealthSheets(workBook) ##Remove older values for Health report

    #logging.basicConfig(filename='c:/asrawat/test/\healthReportGenerator/JiraQueries_Excuted.log', level=logging.WARNING, format='%(asctime)s - %(message)s')
    jiraCredentialFile = open('C:/ASRAWAT/test/JiraAccess.txt', 'r')
    user = jiraCredentialFile.readline().strip()
    password = jiraCredentialFile.readline().strip()
    print(user, password)
    options = {'server': "https://jiradc2.ext.net.nokia.com/"}
    jira = JIRA(options, basic_auth=(user, password))
    issue = jira.issue('RGSOL-3237')
    print(issue.fields.summary)

    queryList = RO.readQueries(querySheet)
    listSize = len(queryList)
    blockedIssueCount = 0
    for queryIndex in range(0, listSize):
        ### Release R20
        release = "R20.8"
        query = queryList[queryIndex] + ' and affectedVersion in ("Nokia Registers 20.8") and sprint in (openSprints())  '
        #print(query)
        # issueList = jira.search_issues(query)
        # if(queryIndex<4):
        #     perfUstCount = perfUstCount + len(issueList)
        #
        # if(queryIndex>3 and queryIndex <10):
        #     servcUstCount = servcUstCount + len(issueList)
        #
        # if(queryIndex> 9 and queryIndex <14):
        #     securityUstcount = securityUstcount + len(issueList)
        #
        # if(queryIndex> 13 and queryIndex <18):
        #     RET_MAVUstCount = RET_MAVUstCount + len(issueList)
        #
        # if(queryIndex> 17 and queryIndex <23):
        #     NetActUstCount = NetActUstCount + len(issueList)
        #
        # if(queryIndex> 23 and queryIndex <30):
        #     fiveGUstCount = fiveGUstCount + len(issueList)

        generateBasicHealthReportReleaseWise(workBook, jira, query, queryIndex, release)
        #Release R19 MP6
        #release = "R19MP6"
        #query = queryList[queryIndex] + ' and affectedVersion in ("Nokia Registers 19MP6") and Sprint = 82868 '
        #generateBasicHealthReportReleaseWise(workBook, jira, query, queryIndex, release)

    fillJiraRaisedBySyVeinReport(workBook, jira) # consolidated Jiras in a single sheet.


    #print("Total blocked USTs = ", blockedIssueCount)

    ## This temporary holder for values
# ## This is to copy values to excel sheet as these variables not available to other module.
#     weeklyTrendSheet.cell(20, 4).value = perfUstCount
#     weeklyTrendSheet.cell(20, 5).value = servcUstCount
#     weeklyTrendSheet.cell(20, 6).value = securityUstcount
#     weeklyTrendSheet.cell(20, 7).value = RET_MAVUstCount
#     weeklyTrendSheet.cell(20, 8).value = NetActUstCount
#     weeklyTrendSheet.cell(20, 9).value = fiveGUstCount
#
#     RO.generateWeeklyTrend(jira, weeklyTrendSheet,chartSheet)
#
#     weeklyTrendSheet.cell(20, 4).value = None
#     weeklyTrendSheet.cell(20, 5).value = None
#     weeklyTrendSheet.cell(20, 6).value = None
#     weeklyTrendSheet.cell(20, 7).value = None
#     weeklyTrendSheet.cell(20, 8).value = None
#     weeklyTrendSheet.cell(20, 9).value = None

    if(guiOrWeb == 0):
        try:
            workBook.save('C:/ASRAWAT/test/BasicHealthReport/BasicHealthForDay2DayWork_wk'+ str(currentWk) +'.xlsx')
        except:
            print("Report is open so saving in temp file tempReport.xlsx")
            workBook.save('C:/ASRAWAT/test/BasicHealthReport/tempReport.xlsx')

        print("Basic Health Report Generated")
        messagebox.showinfo("Title", "Hurray, Health Report Generated.")
        return None
    elif(guiOrWeb == 1):##From Web application
        workBook.save('C:/ASRAWAT/test/Docker/GUI/template/BasicHealthForDay2DayWork.xlsx')
        print("Basic Health Report Generated")
        return workBook


